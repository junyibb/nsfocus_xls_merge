import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
import tkinter as tk
from tkinter import filedialog, messagebox

def add_border(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.border = border

def merge_adjacent_cells(ws):
    last_row = ws.max_row
    
    i = 1
    while i <= last_row:
        current_ip = ws.cell(i, 1).value
        start_row = i

        while i <= last_row and ws.cell(i, 1).value == current_ip:
            if (ws.cell(i, 2).value is None and 
                ws.cell(i, 3).value is None and 
                ws.cell(i, 4).value is None):
                
                end_row = i
                while (i <= last_row and 
                       ws.cell(i, 1).value == current_ip and 
                       ws.cell(i, 2).value is None and 
                       ws.cell(i, 3).value is None and 
                       ws.cell(i, 4).value is None):
                    i += 1
                
                end_row = i - 1

                if start_row != end_row:
                    for col in range(2, 5):
                        if ws.cell(start_row, col).value is not None:
                            ws.merge_cells(start_row=start_row, start_column=col, 
                                           end_row=end_row, end_column=col)
                            merged_cell = ws.cell(start_row, col)
                            merged_cell.alignment = copy(ws.cell(start_row, col).alignment)
            start_row = i
            i += 1

def process_xls_files(directory):
    xls_files = [file for file in os.listdir(directory) if file.endswith('.xls')]
    
    result_df = pd.DataFrame()
    headers_added = False

    for file in xls_files:
        file_path = os.path.join(directory, file)
        try:
            host_overview_df = pd.read_excel(file_path, sheet_name='主机概况', header=None)
            new_column_value = host_overview_df.iloc[2, 1]
        except Exception as e:
            print(f"Error reading '主机概况' sheet in {file}: {e}")
            continue
        
        try:
            remote_vulnerability_df = pd.read_excel(file_path, sheet_name='远程漏洞')
            remote_vulnerability_df.insert(0, '新列', new_column_value)
        except Exception as e:
            print(f"Error reading '远程漏洞' sheet in {file}: {e}")
            continue
        
        if not headers_added:
            result_df = pd.concat([result_df, remote_vulnerability_df], ignore_index=True)
            headers_added = True
        else:
            result_df = pd.concat([result_df, remote_vulnerability_df.iloc[1:]], ignore_index=True)

    result_file = os.path.join(directory, 'result.xlsx')
    result_df.to_excel(result_file, index=False)

    wb = load_workbook(result_file)
    ws = wb.active

    # 修改A2的内容为“IP”
    ws['A2'] = 'IP'

    # 删除第一行
    ws.delete_rows(1)

    # 添加边框和合并单元格
    add_border(ws, f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
    merge_adjacent_cells(ws)
    
    wb.save(result_file)
    messagebox.showinfo("完成", "文件合并成功！结果保存在: " + result_file)

def select_directory():
    directory = filedialog.askdirectory()
    if directory:
        directory_var.set(directory)  # 更新标签显示选定的目录

def start_merge():
    directory = directory_var.get()
    if directory:
        process_xls_files(directory)
    else:
        messagebox.showwarning("警告", "请先选择一个目录！")

def create_gui():
    root = tk.Tk()
    root.title("绿盟xls文档一键合并v1.2")
    root.geometry("400x250")

    # 添加标签
    tk.Label(root, text="微信公众号：毅心安全", font=("Arial", 12)).pack(pady=10)

    # 选择目录按钮
    tk.Button(root, text="选择目录", command=select_directory, font=("Arial", 12)).pack(pady=10)

    # 显示选定目录路径的标签
    global directory_var
    directory_var = tk.StringVar()
    tk.Label(root, textvariable=directory_var, font=("Arial", 10), fg="green").pack(pady=5)

    # “一键合并”按钮
    tk.Button(root, text="一键合并", command=start_merge, font=("Arial", 12)).pack(pady=20)

    # 显示GUI
    root.mainloop()

if __name__ == "__main__":
    create_gui()
